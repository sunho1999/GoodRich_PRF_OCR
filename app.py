#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from flask import Flask, render_template, request, jsonify, session, send_file
from flask_socketio import SocketIO, emit
import os
import json
import tempfile
import threading
import time
import io
from datetime import datetime
import uuid
import logging
from werkzeug.utils import secure_filename

# ê¸°ì¡´ ëª¨ë“ˆë“¤ import
from parsing.pdf_text import PDFTextExtractor
from llm.gpt_summarizer import GPTSummarizer
from core.config import settings

# PDF ìƒì„± ë¼ì´ë¸ŒëŸ¬ë¦¬
try:
    from weasyprint import HTML, CSS
    from markdown2 import markdown
    PDF_GENERATION_AVAILABLE = True
    logger = logging.getLogger(__name__)
    logger.info("âœ… PDF ìƒì„± ë¼ì´ë¸ŒëŸ¬ë¦¬ ë¡œë“œ ì„±ê³µ")
except ImportError as e:
    PDF_GENERATION_AVAILABLE = False
    logger = logging.getLogger(__name__)
    logger.warning(f"âš ï¸ PDF ìƒì„± ë¼ì´ë¸ŒëŸ¬ë¦¬ ë¡œë“œ ì‹¤íŒ¨: {e}")

# Flask ì•± ì´ˆê¸°í™”
app = Flask(__name__)
app.config['SECRET_KEY'] = 'pdf_ocr_analysis_secret_key_2024'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB ì œí•œ

# SocketIO ì´ˆê¸°í™” (ì‹¤ì‹œê°„ ì±„íŒ…ìš©)
socketio = SocketIO(app, cors_allowed_origins="*")

# ë¡œê¹… ì„¤ì •
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ì—…ë¡œë“œ í´ë” ì„¤ì •
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf'}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

class WebAnalyzer:
    """ì›¹ ê¸°ë°˜ PDF ë¶„ì„ê¸°"""
    
    def __init__(self):
        self.config = settings
        self.pdf_extractor = PDFTextExtractor()
        
        # ì‚¬ìš©ìë³„ ë¶„ì„ ë°ì´í„° ì €ì¥ (ë©”ëª¨ë¦¬ ê¸°ë°˜)
        self.user_data = {}
        
        # GPT ì´ˆê¸°í™” (ë” ê°•ë ¥í•œ ì—ëŸ¬ ì²˜ë¦¬)
        try:
            # .env íŒŒì¼ì—ì„œ API í‚¤ ìš°ì„  ë¡œë“œ
            from dotenv import load_dotenv
            load_dotenv()  # .env íŒŒì¼ ê°•ì œ ë¡œë“œ
            
            # .env íŒŒì¼ -> config -> ì‹œìŠ¤í…œ í™˜ê²½ë³€ìˆ˜ ìˆœì„œë¡œ ìš°ì„ ìˆœìœ„
            api_key = self.config.openai_api_key or os.getenv('OPENAI_API_KEY')
            logger.info(f"API í‚¤ í™•ì¸: {'ìˆìŒ' if api_key else 'ì—†ìŒ'}")
            
            if not api_key:
                raise ValueError("OPENAI_API_KEYê°€ .env íŒŒì¼ì— ì„¤ì •ë˜ì§€ ì•ŠìŒ")
            
            self.gpt_summarizer = GPTSummarizer(api_key=api_key)
            self.gpt_available = True
            logger.info("âœ… GPT API ì´ˆê¸°í™” ì„±ê³µ")
        except Exception as e:
            self.gpt_summarizer = None
            self.gpt_available = False
            logger.error(f"âŒ GPT API ì´ˆê¸°í™” ì‹¤íŒ¨: {e}")
            logger.info("ğŸ”„ ê¸°ë³¸ í…ìŠ¤íŠ¸ íŒŒì‹± ëª¨ë“œë¡œ ë™ì‘í•©ë‹ˆë‹¤")
    
    def allowed_file(self, filename):
        """í—ˆìš©ëœ íŒŒì¼ í™•ì¥ì í™•ì¸"""
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    
    def get_user_id(self, request):
        """ì‚¬ìš©ì ID ìƒì„± (IP + User-Agent ê¸°ë°˜)"""
        import hashlib
        user_info = f"{request.remote_addr}_{request.headers.get('User-Agent', '')}"
        return hashlib.md5(user_info.encode()).hexdigest()[:16]
    
    def get_user_data(self, user_id):
        """ì‚¬ìš©ì ë°ì´í„° ê°€ì ¸ì˜¤ê¸°"""
        if user_id not in self.user_data:
            self.user_data[user_id] = {
                'analyzed_products': [],
                'last_activity': datetime.now()
            }
        return self.user_data[user_id]
    
    def save_analysis_result(self, user_id, product_info):
        """ë¶„ì„ ê²°ê³¼ ì €ì¥"""
        user_data = self.get_user_data(user_id)
        user_data['analyzed_products'].append(product_info)
        user_data['last_activity'] = datetime.now()
        
        # ìµœëŒ€ 5ê°œê¹Œì§€ë§Œ ì €ì¥
        if len(user_data['analyzed_products']) > 5:
            user_data['analyzed_products'] = user_data['analyzed_products'][-5:]
        
        logger.info(f"ì‚¬ìš©ì {user_id}ì˜ ë¶„ì„ ê²°ê³¼ ì €ì¥ë¨: {product_info['name']}")
    
    def get_analyzed_products(self, user_id):
        """ë¶„ì„ëœ ìƒí’ˆ ëª©ë¡ ê°€ì ¸ì˜¤ê¸°"""
        user_data = self.get_user_data(user_id)
        return user_data['analyzed_products']
    
    def cleanup_old_data(self):
        """ì˜¤ë˜ëœ ì‚¬ìš©ì ë°ì´í„° ì •ë¦¬ (1ì‹œê°„ ì´ìƒ)"""
        from datetime import timedelta
        cutoff_time = datetime.now() - timedelta(hours=1)
        
        old_users = [
            user_id for user_id, data in self.user_data.items()
            if data['last_activity'] < cutoff_time
        ]
        
        for user_id in old_users:
            del self.user_data[user_id]
            
        if old_users:
            logger.info(f"ì˜¤ë˜ëœ ì‚¬ìš©ì ë°ì´í„° ì •ë¦¬: {len(old_users)}ëª…")
    
    def extract_pdf_content(self, file_path_or_url, is_url=False, use_ocr=True):
        """PDFì—ì„œ í…ìŠ¤íŠ¸ ì¶”ì¶œ (OCR í–¥ìƒ í¬í•¨)"""
        try:
            if is_url:
                success, pages = self.pdf_extractor.extract_text_from_url(file_path_or_url)
            else:
                success, pages = self.pdf_extractor.extract_text_from_pdf(file_path_or_url, use_ocr=use_ocr)
            
            if not success:
                return {
                    'success': False, 
                    'error': 'PDF í…ìŠ¤íŠ¸ ì¶”ì¶œì— ì‹¤íŒ¨í–ˆìŠµë‹ˆë‹¤.'
                }
            
            # ê¸°ë³¸ í…ìŠ¤íŠ¸ ê²°í•© ë° ì¶”ì¶œ í†µê³„ ìƒì„±
            combined_text = ""
            extraction_stats = {
                'total_pages': len(pages),
                'pages_with_text': 0,
                'ocr_enhanced_pages': 0,
                'hybrid_pages': 0,
                'extraction_methods': {}
            }
            
            for page in pages:
                if isinstance(page, dict):
                    text = page.get('text', '')
                    page_num = page.get('page_number', 0)
                    extraction_method = page.get('extraction_method', 'unknown')
                    
                    # í…ìŠ¤íŠ¸ ê²°í•©
                    if text.strip():
                        extraction_stats['pages_with_text'] += 1
                        combined_text += f"\n--- í˜ì´ì§€ {page_num} ---\n{text}\n"
                    
                    # ì¶”ì¶œ ë°©ë²• í†µê³„
                    if extraction_method in extraction_stats['extraction_methods']:
                        extraction_stats['extraction_methods'][extraction_method] += 1
                    else:
                        extraction_stats['extraction_methods'][extraction_method] = 1
                    
                    # OCR í†µê³„
                    if page.get('has_ocr', False):
                        extraction_stats['ocr_enhanced_pages'] += 1
                    if extraction_method == 'hybrid':
                        extraction_stats['hybrid_pages'] += 1
            
            return {
                'success': True,
                'pages': pages,
                'content': combined_text,
                'page_count': len(pages),
                'extraction_stats': extraction_stats
            }
            
        except Exception as e:
            logger.error(f"PDF ì¶”ì¶œ ì˜¤ë¥˜: {e}")
            return {
                'success': False,
                'error': f'PDF ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'
            }
    
    def analyze_product_detail(self, pages, file_name):
        """ìƒí’ˆ ìƒì„¸ ë¶„ì„"""
        if not self.gpt_available:
            return "âŒ GPT APIë¥¼ ì‚¬ìš©í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤. ê¸°ë³¸ í…ìŠ¤íŠ¸ë§Œ ì œê³µë©ë‹ˆë‹¤."
        
        try:
            return self.gpt_summarizer.analyze_for_detail(pages, file_name)
        except Exception as e:
            logger.error(f"GPT ìƒì„¸ ë¶„ì„ ì˜¤ë¥˜: {e}")
            return f"âŒ ë¶„ì„ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}"
    
    def analyze_product_comparison(self, pages, file_name):
        """ìƒí’ˆ ë¹„êµ ë¶„ì„"""
        if not self.gpt_available:
            return "âŒ GPT APIë¥¼ ì‚¬ìš©í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤. ê¸°ë³¸ í…ìŠ¤íŠ¸ë§Œ ì œê³µë©ë‹ˆë‹¤."
        
        try:
            return self.gpt_summarizer.analyze_for_comparison(pages, file_name)
        except Exception as e:
            logger.error(f"GPT ë¹„êµ ë¶„ì„ ì˜¤ë¥˜: {e}")
            return f"âŒ ë¶„ì„ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}"
    
    def generate_chatbot_response(self, question, context):
        """ì±—ë´‡ ì‘ë‹µ ìƒì„±"""
        if not self.gpt_available:
            return "ì£„ì†¡í•©ë‹ˆë‹¤. GPT APIë¥¼ ì‚¬ìš©í•  ìˆ˜ ì—†ì–´ AI ìƒë‹´ ê¸°ëŠ¥ì´ ì œí•œë©ë‹ˆë‹¤."
        
        try:
            prompt = f"""
ë‹¹ì‹ ì€ ì „ë¬¸ ë³´í—˜ ìƒë‹´ì‚¬ì…ë‹ˆë‹¤. ë¶„ì„ëœ ìƒí’ˆ ì •ë³´ë¥¼ ë°”íƒ•ìœ¼ë¡œ ì •í™•í•˜ê³  êµ¬ì²´ì ì¸ ë‹µë³€ì„ ì œê³µí•´ì£¼ì„¸ìš”.

ğŸš¨ **ì¤‘ìš” ì§€ì¹¨**:
1. **í•´ì•½í™˜ê¸‰ê¸ˆ ê´€ë ¨ ì§ˆë¬¸**: ë°˜ë“œì‹œ êµ¬ì²´ì ì¸ ê¸ˆì•¡ê³¼ ë¹„ìœ¨ì„ ì œê³µí•˜ì„¸ìš”
   - ì˜ˆ: "20ë…„(56ì„¸): 6,149,393ì› (29.8%)"
   - ì¼ë°˜ì ì¸ ì„¤ëª…(50% ì§€ê¸‰í˜•)ë³´ë‹¤ëŠ” ì‹¤ì œ í‘œ ë°ì´í„°ë¥¼ ìš°ì„  ì‚¬ìš©

2. **ê°€ì…ë‹´ë³´ ê´€ë ¨ ì§ˆë¬¸**: ì •í™•í•œ ë‹´ë³´ ì •ë³´ë¥¼ ì œê³µí•˜ì„¸ìš”
   - ë‹´ë³´ëª…, ê°€ì…ê¸ˆì•¡, ë³´í—˜ë£Œ, ë§Œê¸°/ë‚©ê¸°ë¥¼ ì •í™•íˆ í‘œì‹œ
   - ì˜ˆ: "ìœ ë°©ì•”(ìˆ˜ìš©ì²´íƒ€ì…)ì§„ë‹¨ë¹„: 4,000ë§Œì›, ë³´í—˜ë£Œ ì—†ìŒ, 100ì„¸ë§Œê¸°/20ë…„ë‚©"
   - ì„¸ë¶€ ë‹´ë³´ê°€ ìˆëŠ” ê²½ìš° ëª¨ë‘ í¬í•¨í•˜ì—¬ ì„¤ëª…

3. **ë³´í—˜ë£Œ ê´€ë ¨ ì§ˆë¬¸**: ì •í™•í•œ ê¸ˆì•¡ì„ ì›ë¬¸ ê·¸ëŒ€ë¡œ ì œê³µ
   - ë°˜ì˜¬ë¦¼í•˜ì§€ ë§ê³  ì›ë³¸ ë¬¸ì„œì˜ ì •í™•í•œ ê°’ì„ ì‚¬ìš©

4. **ë¹„êµ ì§ˆë¬¸**: ë‘ ìƒí’ˆì˜ êµ¬ì²´ì ì¸ ìˆ˜ì¹˜ë¥¼ ë¹„êµí•˜ì—¬ ë‹µë³€

5. **í¸í–¥ ê¸ˆì§€**: ê°ê´€ì ì´ê³  ì¤‘ë¦½ì ì¸ ì •ë³´ë§Œ ì œê³µ
   - íŠ¹ì • ìƒí’ˆì„ ì¶”ì²œí•˜ê±°ë‚˜ ë¹„íŒí•˜ì§€ ë§ê³  ì‚¬ì‹¤ë§Œ ì „ë‹¬
   - ê³ ê°ì´ ìŠ¤ìŠ¤ë¡œ íŒë‹¨í•  ìˆ˜ ìˆë„ë¡ ì •ë³´ ì œê³µ

6. **ë¬¸ì„œ ê¸°ë°˜ ë‹µë³€**: 
   - ë³´í—˜ìƒí’ˆ ì™¸, ì§ˆë¬¸ì€ ì•Œê³ ìˆëŠ” ëŒ€ë¡œ ë‹µë³€
   - ë³´í—˜ê³¼ ê´€ë ¨ì´ ìˆì§€ë§Œ, ë¬¸ì„œì— ì—†ëŠ” ì •ë³´ëŠ” "ë¬¸ì„œì— í•´ë‹¹ ì •ë³´ê°€ ì—†ìŠµë‹ˆë‹¤"ë¼ê³  ëª…ì‹œ
   - ì¶”ì¸¡ì´ë‚˜ ì¼ë°˜ì ì¸ ì •ë³´ ì œê³µ ê¸ˆì§€

**ë¶„ì„ëœ ìƒí’ˆ ì •ë³´**:
{context}

**ê³ ê° ì§ˆë¬¸**: {question}

**ë‹µë³€**: ë¬¸ì„œì˜ ì •í™•í•œ ì •ë³´ë¥¼ ë°”íƒ•ìœ¼ë¡œ êµ¬ì²´ì ì´ê³  ê°ê´€ì ìœ¼ë¡œ ë‹µë³€í•´ì£¼ì„¸ìš”."""
            
            messages = [
                {"role": "system", "content": "ë‹¹ì‹ ì€ ì „ë¬¸ì ì´ê³  ì¹œê·¼í•œ ë³´í—˜ ìƒë‹´ì‚¬ì…ë‹ˆë‹¤. ğŸš¨ ì¤‘ìš”: 1) í•´ì•½í™˜ê¸‰ê¸ˆ í‘œ ë°ì´í„°ë¥¼ ìš°ì„  í™œìš©í•˜ì—¬ êµ¬ì²´ì ì¸ ê¸ˆì•¡ê³¼ ë¹„ìœ¨ ì œê³µ, 2) ê°€ì…ë‹´ë³´ ë¦¬ìŠ¤íŠ¸ì˜ ì •í™•í•œ ì •ë³´ ì œê³µ (ë‹´ë³´ëª…, ê°€ì…ê¸ˆì•¡, ë³´í—˜ë£Œ, ë§Œê¸°/ë‚©ê¸°), 3) ë¬¸ì„œì— ì—†ëŠ” ì •ë³´ëŠ” ì¶”ì¸¡í•˜ì§€ ë§ê³  'ë¬¸ì„œì— í•´ë‹¹ ì •ë³´ê°€ ì—†ìŠµë‹ˆë‹¤'ë¼ê³  ëª…ì‹œ, 4) ê°ê´€ì ì´ê³  ì¤‘ë¦½ì ì¸ ì •ë³´ë§Œ ì œê³µí•˜ì—¬ ê³ ê°ì´ ìŠ¤ìŠ¤ë¡œ íŒë‹¨í•  ìˆ˜ ìˆë„ë¡ ë„ì›€."},
                {"role": "user", "content": prompt}
            ]
            
            response = self.gpt_summarizer._safe_api_call(
                messages=messages, 
                max_tokens=2000,  # í† í° ì œí•œ ì¦ê°€
                retries=2, 
                delay=1
            )
            
            if response is None:
                return "ì£„ì†¡í•©ë‹ˆë‹¤. í˜„ì¬ AI ì‘ë‹µ ìƒì„±ì— ë¬¸ì œê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤. ì ì‹œ í›„ ë‹¤ì‹œ ì‹œë„í•´ì£¼ì„¸ìš”."
            
            return response.choices[0].message.content.strip()
            
        except Exception as e:
            logger.error(f"ì±—ë´‡ ì‘ë‹µ ìƒì„± ì˜¤ë¥˜: {e}")
            return f"ì£„ì†¡í•©ë‹ˆë‹¤. ì‘ë‹µ ìƒì„± ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}"

# ì „ì—­ ë¶„ì„ê¸° ì¸ìŠ¤í„´ìŠ¤
analyzer = WebAnalyzer()

@app.route('/')
def index():
    """ë©”ì¸ í˜ì´ì§€"""
    return render_template('index.html', gpt_available=analyzer.gpt_available)

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """íŒŒì¼ ì—…ë¡œë“œ API"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'íŒŒì¼ì´ ì„ íƒë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤.'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'íŒŒì¼ì´ ì„ íƒë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤.'})
        
        if not analyzer.allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'PDF íŒŒì¼ë§Œ ì—…ë¡œë“œ ê°€ëŠ¥í•©ë‹ˆë‹¤.'})
        
        # ì•ˆì „í•œ íŒŒì¼ëª…ìœ¼ë¡œ ì €ì¥
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_suffix = uuid.uuid4().hex[:8]
        filename = f"{timestamp}_{unique_suffix}_{filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        file.save(file_path)
        
        return jsonify({
            'success': True, 
            'file_path': file_path,
            'filename': filename
        })
        
    except Exception as e:
        logger.error(f"íŒŒì¼ ì—…ë¡œë“œ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'error': f'ì—…ë¡œë“œ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'})

@app.route('/api/analyze/individual', methods=['POST'])
def analyze_individual():
    """ê°œë³„ ìƒí’ˆ ë¶„ì„ API"""
    try:
        data = request.get_json()
        source_type = data.get('source_type')  # 'file' or 'url'
        source = data.get('source')
        product_name = data.get('product_name', 'ìƒí’ˆ')
        
        if not source:
            return jsonify({'success': False, 'error': 'ë¶„ì„í•  ì†ŒìŠ¤ê°€ ì œê³µë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤.'})
        
        # PDF í…ìŠ¤íŠ¸ ì¶”ì¶œ
        is_url = (source_type == 'url')
        result = analyzer.extract_pdf_content(source, is_url=is_url)
        
        if not result['success']:
            return jsonify(result)
        
        # GPT ìƒì„¸ ë¶„ì„ (ê°€ëŠ¥í•œ ê²½ìš°)
        analysis = ""
        gpt_analysis_success = False
        
        if analyzer.gpt_available:
            analysis = analyzer.analyze_product_detail(result['pages'], product_name)
            # GPT ë¶„ì„ì´ ì‹¤ì œë¡œ ì„±ê³µí–ˆëŠ”ì§€ í™•ì¸
            gpt_analysis_success = (
                analysis and 
                not analysis.startswith("âŒ") and 
                not "ë¶„ì„ ì¤‘ ì˜¤ë¥˜ ë°œìƒ" in analysis and
                not "API í˜¸ì¶œì— ì‹¤íŒ¨" in analysis
            )
            logger.info(f"GPT ë¶„ì„ ì„±ê³µ: {gpt_analysis_success}")
        
        # ë©”ëª¨ë¦¬ì— ë¶„ì„ ê²°ê³¼ ì €ì¥ (ì±—ë´‡ìš©)
        user_id = analyzer.get_user_id(request)
        analyzer.save_analysis_result(user_id, {
            'name': product_name,
            'content': result['content'],
            'analysis': analysis,
            'timestamp': datetime.now().isoformat()
        })
        
        return jsonify({
            'success': True,
            'content': result['content'],
            'analysis': analysis,
            'page_count': result['page_count'],
            'gpt_used': gpt_analysis_success,  # ì‹¤ì œ ì„±ê³µ ì—¬ë¶€ë¡œ ë³€ê²½
            'extraction_stats': result.get('extraction_stats', {})
        })
        
    except Exception as e:
        logger.error(f"ê°œë³„ ë¶„ì„ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'error': f'ë¶„ì„ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'})

@app.route('/api/get_raw_text', methods=['POST'])
def get_raw_text():
    """ì›ë³¸ í…ìŠ¤íŠ¸ ì¶”ì¶œ API (ë””ë²„ê¹…ìš©)"""
    try:
        data = request.get_json()
        source = data.get('source')
        source_type = data.get('source_type', 'file')
        
        if not source:
            return jsonify({'success': False, 'error': 'ì†ŒìŠ¤ê°€ ì œê³µë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤.'})
        
        is_url = (source_type == 'url')
        result = analyzer.extract_pdf_content(source, is_url=is_url)
        
        if not result['success']:
            return jsonify(result)
        
        return jsonify({
            'success': True,
            'raw_text': result['content'],
            'page_count': result['page_count'],
            'extraction_stats': result.get('extraction_stats', {})
        })
        
    except Exception as e:
        logger.error(f"ì›ë³¸ í…ìŠ¤íŠ¸ ì¶”ì¶œ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'error': f'ì›ë³¸ í…ìŠ¤íŠ¸ ì¶”ì¶œ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'})

@app.route('/api/analyze/compare', methods=['POST'])
def analyze_compare():
    """2ê°œ ìƒí’ˆ ë¹„êµ ë¶„ì„ API"""
    try:
        data = request.get_json()
        
        # ì²« ë²ˆì§¸ ìƒí’ˆ
        source1_type = data.get('source1_type')
        source1 = data.get('source1')
        product1_name = data.get('product1_name', 'ìƒí’ˆ A')
        
        # ë‘ ë²ˆì§¸ ìƒí’ˆ
        source2_type = data.get('source2_type')
        source2 = data.get('source2')
        product2_name = data.get('product2_name', 'ìƒí’ˆ B')
        
        # ì‚¬ìš©ì ì •ì˜ í”„ë¡¬í”„íŠ¸
        custom_prompt = data.get('custom_prompt', '').strip()
        required_coverages = data.get('required_coverages', [])
        if required_coverages and isinstance(required_coverages, list):
            required_coverages = [str(item).strip() for item in required_coverages if str(item).strip()]
        else:
            required_coverages = []
        
        if not source1 or not source2:
            return jsonify({'success': False, 'error': 'ë¹„êµí•  ë‘ ê°œì˜ ì†ŒìŠ¤ê°€ ëª¨ë‘ í•„ìš”í•©ë‹ˆë‹¤.'})
        
        # ì²« ë²ˆì§¸ ìƒí’ˆ ë¶„ì„
        result1 = analyzer.extract_pdf_content(source1, is_url=(source1_type == 'url'))
        if not result1['success']:
            return jsonify({'success': False, 'error': f'ì²« ë²ˆì§¸ ìƒí’ˆ ë¶„ì„ ì‹¤íŒ¨: {result1["error"]}'})
        
        # ë‘ ë²ˆì§¸ ìƒí’ˆ ë¶„ì„
        result2 = analyzer.extract_pdf_content(source2, is_url=(source2_type == 'url'))
        if not result2['success']:
            return jsonify({'success': False, 'error': f'ë‘ ë²ˆì§¸ ìƒí’ˆ ë¶„ì„ ì‹¤íŒ¨: {result2["error"]}'})
        
        # ë¹„êµ ë¶„ì„ ìˆ˜í–‰ (Rate Limit ë°©ì§€)
        comparison_analysis = ""
        analysis1 = ""
        analysis2 = ""
        gpt_comparison_success = False
        
        if analyzer.gpt_available:
            logger.info("ğŸ¤– GPT ì¢…í•© ë¹„êµ ë¶„ì„ ì‹œì‘...")
            
            try:
                # ìƒˆë¡œìš´ ì¢…í•© ë¹„êµ ë¶„ì„ í•¨ìˆ˜ ì‚¬ìš© (ì‚¬ìš©ì ì •ì˜ í”„ë¡¬í”„íŠ¸ í¬í•¨)
                comparison_analysis = analyzer.gpt_summarizer.analyze_products_comparison(
                    result1['pages'], product1_name,
                    result2['pages'], product2_name,
                    custom_prompt=custom_prompt,
                    required_coverages=required_coverages
                )
                logger.info(f"ğŸ“Š ì¢…í•© ë¹„êµ ë¶„ì„ ê²°ê³¼ ê¸¸ì´: {len(comparison_analysis) if comparison_analysis else 0}")
            except Exception as e:
                logger.error(f"âŒ ì¢…í•© ë¹„êµ ë¶„ì„ ì¤‘ ì˜¤ë¥˜: {e}")
                comparison_analysis = f"âŒ ì¢…í•© ë¹„êµ ë¶„ì„ ì˜¤ë¥˜: {str(e)}"
            
            # ë¶„ì„ ì„±ê³µ ì—¬ë¶€ í™•ì¸
            gpt_comparison_success = (
                comparison_analysis and 
                not comparison_analysis.startswith("âŒ") and 
                not "ë¶„ì„ ì¤‘ ì˜¤ë¥˜ ë°œìƒ" in comparison_analysis and
                not "API í˜¸ì¶œì— ì‹¤íŒ¨" in comparison_analysis
            )
            
            logger.info(f"GPT ì¢…í•© ë¹„êµ ë¶„ì„ ì„±ê³µ: {gpt_comparison_success}")
            
            if gpt_comparison_success:
                logger.info("âœ… ì¢…í•© ë¹„êµ ë¶„ì„ ì™„ë£Œ")
            else:
                logger.warning("âš ï¸ GPT ì¢…í•© ë¹„êµ ë¶„ì„ ì‹¤íŒ¨, ê¸°ë³¸ í…ìŠ¤íŠ¸ ì‚¬ìš©")
                
            # ê°œë³„ ë¶„ì„ë„ ë©”ëª¨ë¦¬ ì €ì¥ìš©ìœ¼ë¡œ ìˆ˜í–‰ (ì„±ê³µí•œ ê²½ìš°ì—ë§Œ)
            if gpt_comparison_success:
                logger.info("ğŸ”„ ê°œë³„ ë¶„ì„ ì‹œì‘ (ë©”ëª¨ë¦¬ ì €ì¥ìš©)...")
                analysis1 = analyzer.analyze_product_comparison(result1['pages'], product1_name)
                time.sleep(1)  # Rate Limit ë°©ì§€
                analysis2 = analyzer.analyze_product_comparison(result2['pages'], product2_name)
            else:
                analysis1 = ""
                analysis2 = ""
        
        # ë©”ëª¨ë¦¬ì— ë¶„ì„ ê²°ê³¼ ì €ì¥
        user_id = analyzer.get_user_id(request)
        analyzer.save_analysis_result(user_id, {
            'name': product1_name,
            'content': result1['content'],
            'analysis': analysis1 if analyzer.gpt_available else "",
            'timestamp': datetime.now().isoformat()
        })
        analyzer.save_analysis_result(user_id, {
            'name': product2_name,
            'content': result2['content'],
            'analysis': analysis2 if analyzer.gpt_available else "",
            'timestamp': datetime.now().isoformat()
        })
        
        return jsonify({
            'success': True,
            'product1': {
                'name': product1_name,
                'content': result1['content'],
                'page_count': result1['page_count']
            },
            'product2': {
                'name': product2_name,
                'content': result2['content'],
                'page_count': result2['page_count']
            },
            'comparison_analysis': comparison_analysis,
            'gpt_used': gpt_comparison_success  # ì‹¤ì œ ì„±ê³µ ì—¬ë¶€ë¡œ ë³€ê²½
        })
        
    except Exception as e:
        logger.error(f"ë¹„êµ ë¶„ì„ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'error': f'ë¹„êµ ë¶„ì„ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'})

# WebSocket ì´ë²¤íŠ¸ (ì±—ë´‡ìš©)
@socketio.on('connect')
def handle_connect():
    """í´ë¼ì´ì–¸íŠ¸ ì—°ê²°"""
    logger.info(f"í´ë¼ì´ì–¸íŠ¸ ì—°ê²°: {request.sid}")
    emit('status', {'message': 'AI ìƒë‹´ ì„œë¹„ìŠ¤ì— ì—°ê²°ë˜ì—ˆìŠµë‹ˆë‹¤.', 'type': 'info'})

@socketio.on('disconnect')
def handle_disconnect():
    """í´ë¼ì´ì–¸íŠ¸ ì—°ê²° í•´ì œ"""
    logger.info(f"í´ë¼ì´ì–¸íŠ¸ ì—°ê²° í•´ì œ: {request.sid}")

@socketio.on('chat_message')
def handle_chat_message(data):
    """ì±„íŒ… ë©”ì‹œì§€ ì²˜ë¦¬"""
    try:
        question = data.get('message', '').strip()
        if not question:
            emit('chat_response', {'error': 'ë©”ì‹œì§€ë¥¼ ì…ë ¥í•´ì£¼ì„¸ìš”.'})
            return
        
        # ë¶„ì„ëœ ìƒí’ˆ ì •ë³´ ê°€ì ¸ì˜¤ê¸° (ë©”ëª¨ë¦¬ì—ì„œ)
        from flask import request as flask_request
        user_id = analyzer.get_user_id(flask_request)
        analyzed_products = analyzer.get_analyzed_products(user_id)
        
        if not analyzed_products:
            emit('chat_response', {
                'response': 'ì•„ì§ ë¶„ì„ëœ ìƒí’ˆì´ ì—†ìŠµë‹ˆë‹¤. ë¨¼ì € ìƒí’ˆì„ ë¶„ì„í•´ì£¼ì„¸ìš”.',
                'type': 'warning'
            })
            return
        
        # ì»¨í…ìŠ¤íŠ¸ êµ¬ì„± (ì „ì²´ ë‚´ìš© í¬í•¨)
        context = "ë¶„ì„ëœ ìƒí’ˆ ì •ë³´:\n\n"
        for i, product in enumerate(analyzed_products[-3:], 1):  # ìµœê·¼ 3ê°œë§Œ ì‚¬ìš©
            # ì „ì²´ ë‚´ìš© í¬í•¨ (1500ì ì œí•œ í•´ì œ)
            context += f"ìƒí’ˆ {i}: {product['name']}\n{product['content']}\n\n"
            
        # ë””ë²„ê¹…ì„ ìœ„í•œ ë¡œê¹…
        logger.info(f"ì±—ë´‡ ì»¨í…ìŠ¤íŠ¸ ê¸¸ì´: {len(context)} ì")
        logger.info(f"ë¶„ì„ëœ ìƒí’ˆ ìˆ˜: {len(analyzed_products)}")
        
        # ë¡œë”© ìƒíƒœ ì „ì†¡
        emit('chat_response', {'loading': True, 'message': 'AIê°€ ì‘ë‹µì„ ìƒì„±í•˜ê³  ìˆìŠµë‹ˆë‹¤...'})
        
        # í˜„ì¬ request sid ì €ì¥
        current_sid = request.sid
        
        # ë³„ë„ ìŠ¤ë ˆë“œì—ì„œ AI ì‘ë‹µ ìƒì„±
        def generate_response():
            try:
                logger.info(f"ì±—ë´‡ ì§ˆë¬¸: {question}")
                logger.info(f"ì»¨í…ìŠ¤íŠ¸ ë¯¸ë¦¬ë³´ê¸°: {context[:500]}...")
                
                response = analyzer.generate_chatbot_response(question, context)
                
                logger.info(f"ì±—ë´‡ ì‘ë‹µ: {response[:200]}...")
                
                socketio.emit('chat_response', {
                    'response': response,
                    'type': 'success',
                    'loading': False
                }, room=current_sid)
            except Exception as e:
                logger.error(f"ì±—ë´‡ ì‘ë‹µ ìƒì„± ì˜¤ë¥˜: {e}")
                socketio.emit('chat_response', {
                    'error': f'ì‘ë‹µ ìƒì„± ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}',
                    'loading': False
                }, room=current_sid)
        
        thread = threading.Thread(target=generate_response)
        thread.daemon = True
        thread.start()
        
    except Exception as e:
        logger.error(f"ì±„íŒ… ì²˜ë¦¬ ì˜¤ë¥˜: {e}")
        emit('chat_response', {'error': f'ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'})

def start_app():
    """ì•± ì‹œì‘ í•¨ìˆ˜ - Vercelê³¼ ë¡œì»¬ í™˜ê²½ ëª¨ë‘ ì§€ì›"""
    logger.info("ğŸš€ PDF OCR ì›¹ ì• í”Œë¦¬ì¼€ì´ì…˜ ì‹œì‘...")
    logger.info(f"GPT API ìƒíƒœ: {'âœ… ì‚¬ìš© ê°€ëŠ¥' if analyzer.gpt_available else 'âŒ ì‚¬ìš© ë¶ˆê°€'}")
    
    # ì£¼ê¸°ì  ë°ì´í„° ì •ë¦¬ ìŠ¤ì¼€ì¤„ë§
    import threading
    import time
    
    def cleanup_task():
        while True:
            time.sleep(300)  # 5ë¶„ë§ˆë‹¤ ì‹¤í–‰
            analyzer.cleanup_old_data()
    
    cleanup_thread = threading.Thread(target=cleanup_task, daemon=True)
    cleanup_thread.start()
    
    # í™˜ê²½ì— ë”°ë¼ ì‹¤í–‰ ë°©ì‹ ê²°ì •
    debug_mode = os.getenv('FLASK_DEBUG', 'False').lower() == 'true'
    
    if debug_mode:
        # ê°œë°œ í™˜ê²½
        socketio.run(app, host='0.0.0.0', port=8080, debug=True)
    else:
        # í”„ë¡œë•ì…˜ í™˜ê²½
        socketio.run(app, host='0.0.0.0', port=8080, debug=False, allow_unsafe_werkzeug=True)

@app.route('/api/generate_pdf', methods=['POST'])
def generate_pdf():
    """ë¶„ì„ ê²°ê³¼ë¥¼ PDFë¡œ ìƒì„±í•˜ì—¬ ë‹¤ìš´ë¡œë“œ"""
    try:
        data = request.get_json()
        markdown_content = data.get('content', '')
        filename = data.get('filename', 'ë³´í—˜ìƒí’ˆ_ë¶„ì„ê²°ê³¼')
        
        if not markdown_content:
            return jsonify({
                'success': False,
                'error': 'ë³€í™˜í•  ë‚´ìš©ì´ ì—†ìŠµë‹ˆë‹¤.'
            }), 400
        
        # PDF ìƒì„± ë¼ì´ë¸ŒëŸ¬ë¦¬ í™•ì¸
        if not PDF_GENERATION_AVAILABLE:
            return jsonify({
                'success': False,
                'error': 'PDF ìƒì„± ê¸°ëŠ¥ì´ ì‚¬ìš© ë¶ˆê°€ëŠ¥í•©ë‹ˆë‹¤. weasyprintì™€ markdown2ë¥¼ ì„¤ì¹˜í•´ì£¼ì„¸ìš”.'
            }), 500
        
        # ë§ˆí¬ë‹¤ìš´ì„ HTMLë¡œ ë³€í™˜
        html_content = markdown(markdown_content, extras=['tables', 'fenced-code-blocks'])
        
        # HTML í…œí”Œë¦¿
        html_template = f"""
        <!DOCTYPE html>
        <html lang="ko">
        <head>
            <meta charset="UTF-8">
            <title>{filename}</title>
            <style>
                @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700&display=swap');
                
                body {{
                    font-family: 'Noto Sans KR', sans-serif;
                    line-height: 1.4;
                    color: #333;
                    margin: 0;
                    padding: 10mm;
                    font-size: 9pt;
                }}
                
                h1 {{
                    color: #2c3e50;
                    text-align: center;
                    margin: 10mm 0 5mm 0;
                    font-size: 18pt;
                    font-weight: 700;
                }}
                
                h2 {{
                    color: #34495e;
                    border-bottom: 2px solid #95a5a6;
                    padding-bottom: 8px;
                    margin-top: 15px;
                    margin-bottom: 10px;
                    font-size: 18pt;
                    font-weight: 600;
                    page-break-after: avoid;
                    page-break-before: auto;
                }}
                
                h3 {{
                    color: #7f8c8d;
                    margin-top: 20px;
                    font-size: 14pt;
                    font-weight: 500;
                }}
                
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 5mm 0;
                    font-size: 7pt;
                }}
                
                /* ë³´ì¥ í•­ëª© ì»¬ëŸ¼ (ì²« ë²ˆì§¸ ì»¬ëŸ¼) */
                th:first-child, td:first-child {{
                    width: 25%;
                    text-align: left;
                    padding-left: 8px;
                    font-weight: 600;
                    background-color: #f8f9fa !important;
                }}
                
                /* ë‚˜ë¨¸ì§€ ì»¬ëŸ¼ */
                th, td {{
                    border: 1px solid #d0d0d0;
                    padding: 6px 4px;
                    text-align: center;
                    vertical-align: middle;
                }}
                
                /* 4ë²ˆì§¸ì™€ 5ë²ˆì§¸ ì»¬ëŸ¼ ì‚¬ì´ êµµì€ ì¤„ */
                th:nth-child(5), td:nth-child(5) {{
                    border-left: 3px solid #8B9FE8;
                }}
                
                /* ë©”ì¸ í—¤ë” */
                thead th {{
                    background: linear-gradient(135deg, #6B7FD7 0%, #8B9FE8 100%);
                    color: white;
                    font-weight: 700;
                    font-size: 8pt;
                    padding: 8px;
                }}
                
                /* ë³´ì¥ í•­ëª© ì»¬ëŸ¼ í—¤ë” */
                thead th:first-child {{
                    background: linear-gradient(135deg, #6B7FD7 0%, #8B9FE8 100%);
                    color: white;
                    font-weight: 700;
                }}
                
                /* êµì°¨ í–‰ ë°°ê²½ìƒ‰ (ë…¸ë€ìƒ‰/íŒŒë€ìƒ‰) */
                /* í™€ìˆ˜ í–‰ - ì „ì²´ ë…¸ë€ìƒ‰ */
                tbody tr:nth-child(odd) td:nth-child(2),
                tbody tr:nth-child(odd) td:nth-child(3),
                tbody tr:nth-child(odd) td:nth-child(4),
                tbody tr:nth-child(odd) td:nth-child(5),
                tbody tr:nth-child(odd) td:nth-child(6),
                tbody tr:nth-child(odd) td:nth-child(7) {{
                    background-color: #FFF9E6;
                }}
                
                /* ì§ìˆ˜ í–‰ - ì „ì²´ íŒŒë€ìƒ‰ */
                tbody tr:nth-child(even) td:nth-child(2),
                tbody tr:nth-child(even) td:nth-child(3),
                tbody tr:nth-child(even) td:nth-child(4),
                tbody tr:nth-child(even) td:nth-child(5),
                tbody tr:nth-child(even) td:nth-child(6),
                tbody tr:nth-child(even) td:nth-child(7) {{
                    background-color: #E8F4FD;
                }}
                
                /* ì‹ ê·œ ë‹´ë³´ ê°•ì¡° */
                .new-coverage {{
                    background-color: #d4edda !important;
                    border-left: 3px solid #28a745 !important;
                }}
                
                hr {{
                    border: none;
                    border-top: 1px solid #ecf0f1;
                    margin: 25px 0;
                }}
                
                ul, ol {{
                    margin: 10px 0;
                    padding-left: 25px;
                }}
                
                li {{
                    margin: 5px 0;
                }}
                
                strong {{
                    color: #2c3e50;
                    font-weight: 600;
                }}
                
                code {{
                    background-color: #f4f4f4;
                    padding: 2px 5px;
                    border-radius: 3px;
                    font-family: 'Courier New', monospace;
                }}
                
                @page {{
                    size: A4;
                    margin: 15mm;
                    @bottom-center {{
                        content: counter(page) " / " counter(pages);
                        font-size: 9pt;
                        color: #7f8c8d;
                    }}
                }}
            </style>
        </head>
        <body>
            {html_content}
        </body>
        </html>
        """
        
        # HTMLì„ PDFë¡œ ë³€í™˜
        pdf_bytes = HTML(string=html_template, base_url='.').write_pdf()
        
        # BytesIO ê°ì²´ë¡œ ë³€í™˜
        pdf_buffer = io.BytesIO(pdf_bytes)
        pdf_buffer.seek(0)
        
        # íŒŒì¼ëª… ìƒì„±
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        pdf_filename = f"{filename}_{timestamp}.pdf"
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=pdf_filename
        )
        
    except Exception as e:
        logger.error(f"PDF ìƒì„± ì˜¤ë¥˜: {e}")
        return jsonify({
            'success': False,
            'error': f'PDF ìƒì„± ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'
        }), 500

@app.route('/api/generate_excel', methods=['POST'])
def generate_excel():
    """ë¶„ì„ ê²°ê³¼ë¥¼ Excelë¡œ ìƒì„±í•˜ì—¬ ë‹¤ìš´ë¡œë“œ"""
    try:
        data = request.get_json()
        markdown_content = data.get('content', '')
        filename = data.get('filename', 'ë³´í—˜ìƒí’ˆ_ë¶„ì„ê²°ê³¼')
        
        if not markdown_content:
            return jsonify({
                'success': False,
                'error': 'ë³€í™˜í•  ë‚´ìš©ì´ ì—†ìŠµë‹ˆë‹¤.'
            }), 400
        
        # Excel ìƒì„± ë¼ì´ë¸ŒëŸ¬ë¦¬ í™•ì¸
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({
                'success': False,
                'error': 'Excel ìƒì„± ê¸°ëŠ¥ì´ ì‚¬ìš© ë¶ˆê°€ëŠ¥í•©ë‹ˆë‹¤. openpyxlì„ ì„¤ì¹˜í•´ì£¼ì„¸ìš”.'
            }), 500
        
        # ì›Œí¬ë¶ ìƒì„±
        wb = Workbook()
        wb.remove(wb.active)  # ê¸°ë³¸ ì‹œíŠ¸ ì œê±°
        
        # ë§ˆí¬ë‹¤ìš´ ë‚´ìš©ì—ì„œ í‘œ ì¶”ì¶œ
        sections = extract_tables_from_markdown(markdown_content)
        
        # í‘œê°€ ì—†ìœ¼ë©´ ì˜¤ë¥˜ ë°˜í™˜
        if not sections or all(not tables for tables in sections.values()):
            return jsonify({
                'success': False,
                'error': 'í‘œ í˜•ì‹ì˜ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤. Excelë¡œ ë³€í™˜í•  ìˆ˜ ìˆëŠ” í‘œê°€ í•„ìš”í•©ë‹ˆë‹¤.'
            }), 400
        
        # ê° ì„¹ì…˜ì„ ë³„ë„ ì‹œíŠ¸ë¡œ ìƒì„±
        for section_name, tables in sections.items():
            if not tables:
                continue
            
            # ì‹œíŠ¸ ìƒì„±
            ws = wb.create_sheet(title=section_name[:31])  # Excel ì‹œíŠ¸ ì´ë¦„ì€ 31ì ì œí•œ
            
            # ìŠ¤íƒ€ì¼ ì •ì˜
            header_fill = PatternFill(start_color="6B7FD7", end_color="8B9FE8", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            row_num = 1
            
            # ê° í‘œë¥¼ ì‹œíŠ¸ì— ì¶”ê°€
            for table_idx, table in enumerate(tables):
                if table_idx > 0:
                    row_num += 2  # í‘œ ì‚¬ì´ ê°„ê²©
                
                # í‘œ ë°ì´í„°ë¥¼ ì‹œíŠ¸ì— ì‘ì„±
                for r_idx, row in enumerate(table):
                    for c_idx, cell_value in enumerate(row):
                        cell = ws.cell(row=row_num + r_idx, column=c_idx + 1, value=cell_value)
                        cell.border = border
                        
                        # í—¤ë” í–‰ ìŠ¤íƒ€ì¼ ì ìš©
                        if r_idx == 0:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        else:
                            # ì²« ë²ˆì§¸ ì—´ì€ ì™¼ìª½ ì •ë ¬, ë‚˜ë¨¸ì§€ëŠ” ê°€ìš´ë° ì •ë ¬
                            if c_idx == 0:
                                cell.alignment = left_alignment
                            else:
                                cell.alignment = center_alignment
                            
                            # êµì°¨ í–‰ ë°°ê²½ìƒ‰ ì ìš©
                            if r_idx % 2 == 0:  # ì§ìˆ˜ í–‰ (0-indexedì´ë¯€ë¡œ ì‹¤ì œë¡œëŠ” í™€ìˆ˜ ë²ˆì§¸ í–‰)
                                cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
                            else:  # í™€ìˆ˜ í–‰ (ì‹¤ì œë¡œëŠ” ì§ìˆ˜ ë²ˆì§¸ í–‰)
                                cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
                
                # ì—´ ë„ˆë¹„ ìë™ ì¡°ì •
                for col_idx in range(len(table[0]) if table else 0):
                    max_length = 0
                    col_letter = get_column_letter(col_idx + 1)
                    
                    for row in table:
                        if col_idx < len(row):
                            cell_value = str(row[col_idx])
                            max_length = max(max_length, len(cell_value))
                    
                    # ìµœì†Œ ë„ˆë¹„ ì„¤ì •
                    adjusted_width = max(max_length + 2, 10)
                    # ìµœëŒ€ ë„ˆë¹„ ì œí•œ (ë„ˆë¬´ ë„“ì§€ ì•Šê²Œ)
                    adjusted_width = min(adjusted_width, 50)
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                row_num += len(table)
        
        # ë©”ëª¨ë¦¬ì— Excel íŒŒì¼ ì €ì¥
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # íŒŒì¼ëª… ìƒì„±
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"{filename}_{timestamp}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=excel_filename
        )
        
    except Exception as e:
        logger.error(f"Excel ìƒì„± ì˜¤ë¥˜: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': f'Excel ìƒì„± ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}'
        }), 500

def extract_tables_from_markdown(markdown_content):
    """
    ë§ˆí¬ë‹¤ìš´ ë‚´ìš©ì—ì„œ í‘œë¥¼ ì¶”ì¶œí•˜ì—¬ ë”•ì…”ë„ˆë¦¬ë¡œ ë°˜í™˜
    ë°˜í™˜ í˜•ì‹: {'ì‹œíŠ¸ëª…': [í‘œ1, í‘œ2, ...]}
    ê° í‘œëŠ” 2ì°¨ì› ë¦¬ìŠ¤íŠ¸: [í—¤ë”í–‰, ë°ì´í„°í–‰1, ë°ì´í„°í–‰2, ...]
    """
    sections = {}
    lines = markdown_content.split('\n')
    
    current_section = None
    current_table = None
    in_table = False
    
    for line in lines:
        line = line.strip()
        
        # ì„¹ì…˜ í—¤ë” ê°ì§€ (## 1. ìš”ì•½ ë¹„êµí‘œ, ## 2. ê³µí†µ ê°€ì…ë‹´ë³´ ë¹„êµ ë“±)
        if line.startswith('##'):
            # í˜„ì¬ í‘œê°€ ìˆìœ¼ë©´ ì €ì¥
            if current_table and current_section:
                if current_section not in sections:
                    sections[current_section] = []
                sections[current_section].append(current_table)
                current_table = None
            
            # ì„¹ì…˜ ì´ë¦„ ì¶”ì¶œ
            section_name = line.replace('##', '').strip()
            # ë²ˆí˜¸ ì œê±° (ì˜ˆ: "1. ìš”ì•½ ë¹„êµí‘œ" -> "ìš”ì•½ ë¹„êµí‘œ")
            section_name = section_name.split('.', 1)[-1].strip() if '.' in section_name else section_name
            current_section = section_name
            in_table = False
            continue
        
        # ë§ˆí¬ë‹¤ìš´ í…Œì´ë¸” í–‰ ê°ì§€
        if line.startswith('|') and line.endswith('|'):
            if not in_table:
                in_table = True
                current_table = []
            
            # ì…€ ì¶”ì¶œ
            cells = [cell.strip() for cell in line.split('|')[1:-1]]  # ì²« ë²ˆì§¸ì™€ ë§ˆì§€ë§‰ ë¹ˆ ìš”ì†Œ ì œê±°
            
            # êµ¬ë¶„ì„  í–‰ ì œì™¸ (--- ë˜ëŠ” :--- ê°™ì€ íŒ¨í„´)
            if all(cell.replace('-', '').replace(':', '').strip() == '' for cell in cells):
                continue
            
            # ë§ˆí¬ë‹¤ìš´ ê°•ì¡° ì œê±° (**í…ìŠ¤íŠ¸** -> í…ìŠ¤íŠ¸)
            cells = [cell.replace('**', '').replace('*', '').strip() for cell in cells]
            
            if cells:
                current_table.append(cells)
        else:
            # í…Œì´ë¸”ì´ ëë‚¨
            if in_table and current_table:
                if current_section:
                    if current_section not in sections:
                        sections[current_section] = []
                    sections[current_section].append(current_table)
                current_table = None
            in_table = False
    
    # ë§ˆì§€ë§‰ í‘œ ì €ì¥
    if current_table and current_section:
        if current_section not in sections:
            sections[current_section] = []
        sections[current_section].append(current_table)
    
    return sections

# Vercel í™˜ê²½ì—ì„œëŠ” importë§Œ ë˜ê³  ì‹¤í–‰ë˜ì§€ ì•ŠìŒ
# ë¡œì»¬ í™˜ê²½ì—ì„œë§Œ ì‹¤í–‰
if __name__ == '__main__':
    start_app()

# Vercelìš© WSGI ì•± export
application = app
